import openpyxl

def getRowCount(file,sheetName):
    Workbook = openpyxl.load_workbook(file)
    sheet = Workbook[sheetName]
    return (sheet.max_row)

def getColumnCount(file,sheetName):
    Workbook = openpyxl.load_workbook(file)
    sheet = Workbook[sheetName]
    return (sheet.max_column)

def readData(file,sheetName,rownum,colnum):
    Workbook = openpyxl.load_workbook(file)
    sheet = Workbook[sheetName]
    return sheet.cell(row=rownum,column=colnum).value

def writeData(file,sheetName,rownum,colnum,data):
    Workbook = openpyxl.load_workbook(file)
    sheet = Workbook[sheetName]
    sheet.cell(row=rownum,column=colnum).value = data
    Workbook.save(file)









